#!/usr/bin/env python3
import argparse
import os
import re
import sys
import time
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright, Download, Page, BrowserContext

# User-Agent moderno y realista de Chrome en Windows.
REALISTIC_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/134.0.0.0 Safari/537.36"
)

DOWNLOAD_REGEX = re.compile(r"(download|descargar)", re.I)
PASSCODE_LABEL_REGEX = re.compile(r"(passcode|codigo|código|clave)", re.I)


def log(msg: str) -> None:
    print(msg, flush=True)


def safe_name(value: Optional[str], fallback: str = "sin_nombre") -> str:
    text = (value or "").strip()
    if not text:
        text = fallback
    # Solo filtrar caracteres inválidos para Windows, pero conservar tildes y especiales
    text = re.sub(r"[\\/:*?\"<>|]+", "_", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text[:180] or fallback


def detect_headers(sheet) -> Dict[str, int]:
    header_map = {}
    for cell in sheet[1]:
        value = str(cell.value).strip().lower() if cell.value is not None else ""
        if value:
            header_map[value] = cell.column

    required = {
        "url": None,
        "codigo": None,
        "clase": None,
        "sesion": None,
        "fecha": None,
    }

    for key in list(header_map.keys()):
        if key == "url":
            required["url"] = header_map[key]
        elif key in {"código", "codigo", "código ", "codigo "}:
            required["codigo"] = header_map[key]
        elif key.startswith("clase"):
            required["clase"] = header_map[key]
        elif key.startswith("sesión") or key.startswith("sesion"):
            required["sesion"] = header_map[key]
        elif key.startswith("fecha"):
            required["fecha"] = header_map[key]

    missing = [k for k in ["url", "codigo"] if required[k] is None]
    if missing:
        raise ValueError(f"No encontré las columnas requeridas: {', '.join(missing)}")

    return required


def load_rows(xlsx_path: Path) -> List[Dict[str, str]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cols = detect_headers(ws)

    rows: List[Dict[str, str]] = []
    current_class = ""

    for row_idx in range(3, ws.max_row + 1):
        clase = ws.cell(row=row_idx, column=cols["clase"]).value if cols["clase"] else None
        sesion = ws.cell(row=row_idx, column=cols["sesion"]).value if cols["sesion"] else None
        fecha = ws.cell(row=row_idx, column=cols["fecha"]).value if cols["fecha"] else None
        url = ws.cell(row=row_idx, column=cols["url"]).value if cols["url"] else None
        codigo = ws.cell(row=row_idx, column=cols["codigo"]).value if cols["codigo"] else None

        if clase:
            current_class = str(clase).strip()

        if not url:
            continue

        rows.append(
            {
                "row": str(row_idx),
                "clase": current_class,
                "sesion": str(sesion or "").strip(),
                "fecha": str(fecha or "").strip(),
                "url": str(url).strip(),
                "codigo": str(codigo or "").strip(),
            }
        )

    return rows


def build_filename_prefix(item: Dict[str, str], index: int) -> str:
    parts = [
        f"{index:04d}",
        safe_name(item.get("clase"), "clase"),
        safe_name(item.get("sesion"), "sesion"),
        safe_name(item.get("fecha"), "fecha"),
    ]
    return " - ".join(parts)


def click_if_visible(locator, timeout: int = 5000) -> bool:
    try:
        locator.first.wait_for(state="visible", timeout=timeout)
        locator.first.click(timeout=timeout)
        return True
    except Exception:
        return False


def resolve_browser_context(playwright, headless: bool, downloads_dir: Path) -> BrowserContext:
    browser = None
    chosen = None

    for channel in ("chrome", "msedge"):
        try:
            browser = playwright.chromium.launch(
                channel=channel,
                headless=headless,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--start-maximized",
                ],
            )
            chosen = channel
            break
        except Exception:
            browser = None

    if browser is None:
        browser = playwright.chromium.launch(
            headless=headless,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--start-maximized",
            ],
        )
        chosen = "chromium"

    log(f"[INFO] Navegador seleccionado: {chosen}")

    context = browser.new_context(
        accept_downloads=True,
        user_agent=REALISTIC_UA,
        viewport={"width": 1440, "height": 900},
        locale="es-ES",
        timezone_id="America/Santiago",
    )

    context.set_default_timeout(20000)
    downloads_dir.mkdir(parents=True, exist_ok=True)
    return context


def try_submit_passcode(page: Page, code: str) -> None:
    if not code:
        return

    candidates = [
        page.locator('input[type="password"]'),
        page.get_by_label(PASSCODE_LABEL_REGEX),
        page.get_by_placeholder(PASSCODE_LABEL_REGEX),
        page.locator('input[aria-label*="passcode" i]'),
        page.locator('input[placeholder*="passcode" i]'),
    ]

    field = None
    for locator in candidates:
        try:
            if locator.first.is_visible(timeout=3000):
                field = locator.first
                break
        except Exception:
            continue

    if field is None:
        return

    field.fill(code)

    submitters = [
        page.get_by_role("button", name=re.compile(r"(view recording|watch recording|submit|continuar|continuar a la grabación|acceder|ver grabación)", re.I)),
        page.locator('button[type="submit"]'),
        page.locator('input[type="submit"]'),
    ]

    for locator in submitters:
        if click_if_visible(locator, timeout=5000):
            try:
                page.wait_for_selector(".download-btn", timeout=20000)
            except Exception:
                page.wait_for_load_state("networkidle", timeout=15000)
            return

    field.press("Enter")
    try:
        page.wait_for_selector(".download-btn", timeout=20000)
    except Exception:
        page.wait_for_load_state("networkidle", timeout=15000)


def dismiss_cookie_or_modal(page: Page) -> None:
    locators = [
        page.get_by_role("button", name=re.compile(r"(accept|aceptar|ok|entendido|got it)", re.I)),
        page.get_by_role("button", name=re.compile(r"(close|cerrar)", re.I)),
    ]
    for locator in locators:
        try:
            if locator.first.is_visible(timeout=1500):
                locator.first.click(timeout=1500)
        except Exception:
            pass


def collect_download_buttons(page: Page):
    selectors = [
        page.get_by_role("button", name=DOWNLOAD_REGEX),
        page.get_by_role("link", name=DOWNLOAD_REGEX),
        page.locator("button", has_text=DOWNLOAD_REGEX),
        page.locator("a", has_text=DOWNLOAD_REGEX),
        page.locator('[data-testid*="download" i]'),
    ]
    seen = []
    for locator in selectors:
        try:
            count = min(locator.count(), 10)
        except Exception:
            count = 0
        for i in range(count):
            seen.append(locator.nth(i))
    return seen


def save_download(download: Download, target_dir: Path, prefix: str, index: int) -> Path:
    suggested = download.suggested_filename
    suffix = Path(suggested).suffix or ".mp4"
    filename = f"{prefix} - archivo_{index:02d}{suffix}"
    target = target_dir / filename
    download.save_as(str(target))
    return target


def process_item(context: BrowserContext, item: Dict[str, str], index: int, out_dir: Path) -> None:
    prefix = build_filename_prefix(item, index)
    target_dir = out_dir / safe_name(item.get("clase"), "clase")
    target_dir.mkdir(parents=True, exist_ok=True)

    page = context.new_page()
    log(f"\n[INFO] Fila {item['row']} -> {item['sesion'] or item['url']}")

    try:
        page.goto(item["url"], wait_until="domcontentloaded", timeout=30000)
        page.wait_for_load_state("networkidle", timeout=15000)
        dismiss_cookie_or_modal(page)
        try_submit_passcode(page, item.get("codigo", ""))
        dismiss_cookie_or_modal(page)

        aria2_links = []
        buttons = collect_download_buttons(page)
        saved = 0
        if buttons:
            for btn in buttons:
                try:
                    # Extraer href si es <a>, sino descargar normalmente
                    href = btn.get_attribute("href") if btn.evaluate("el => el.tagName.toLowerCase() === 'a'") else None
                    if href:
                        aria2_links.append(href)
                    else:
                        with page.expect_download(timeout=15000) as dl_info:
                            btn.click(timeout=5000)
                        download = dl_info.value
                        path = save_download(download, target_dir, prefix, saved + 1)
                        saved += 1
                        log(f"[OK] Descargado: {path}")
                        time.sleep(1)
                except PlaywrightTimeoutError:
                    continue
                except Exception as exc:
                    log(f"[WARN] No se pudo descargar desde un botón: {exc}")
                    continue

        # Si no se descargó nada, buscar todos los enlaces de descarga en el web player
        if saved == 0:
            try:
                download_links = page.locator(".download-btn")
                count = download_links.count()
                if count > 0:
                    for i in range(count):
                        link = download_links.nth(i)
                        if link.is_visible(timeout=5000):
                            href = link.get_attribute("href")
                            if href:
                                aria2_links.append(href)
                            else:
                                with page.expect_download(timeout=15000) as dl_info:
                                    link.click(timeout=5000)
                                download = dl_info.value
                                path = save_download(download, target_dir, prefix, saved + 1)
                                log(f"[OK] Descargado desde web player: {path}")
                                saved += 1
                else:
                    log("[WARN] No se encontraron enlaces de descarga en el web player.")
            except Exception as exc:
                log(f"[WARN] No se pudo descargar desde los enlaces del web player: {exc}")

        # Guardar enlaces para aria2c y ejecutar descargas
        if aria2_links:
            links_file = target_dir / f"{prefix}_aria2c.txt"
            with open(links_file, "w", encoding="utf-8") as f:
                for url in aria2_links:
                    f.write(url + "\n")
            log(f"[INFO] Enlaces guardados para aria2c: {links_file}")

            # Ejecutar aria2c para descargar todos los enlaces
            import subprocess
            aria2_cmd = ["aria2c", "-x", "16", "-j", "8", "-d", str(target_dir), "-i", str(links_file)]
            try:
                log(f"[INFO] Ejecutando aria2c para descargas paralelas...")
                result = subprocess.run(aria2_cmd, capture_output=True, text=True)
                log(result.stdout)
                if result.stderr:
                    log(f"[WARN] aria2c error: {result.stderr}")
            except Exception as exc:
                log(f"[ERROR] No se pudo ejecutar aria2c: {exc}")

        if saved == 0 and not aria2_links:
            raise RuntimeError("La página abrió, pero no entregó ninguna descarga.")

        log(f"[INFO] Última fila procesada: {item['row']}")

    finally:
        page.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Descarga grabaciones de Zoom listadas en una planilla XLSX.")
    parser.add_argument("xlsx", type=Path, help="Ruta del archivo XLSX")
    parser.add_argument("--output", type=Path, default=Path("./descargas_zoom"), help="Carpeta destino")
    parser.add_argument("--headless", action="store_true", help="Ejecuta sin interfaz gráfica")
    parser.add_argument("--start-row", type=int, default=1, help="Índice lógico para retomar (1 = primer registro válido)")
    parser.add_argument("--limit", type=int, default=0, help="Máximo de registros a procesar. 0 = todos")
    parser.add_argument("--aria2c", action="store_true", help="Usa aria2c para descargas paralelas (opcional)")
    parser.add_argument("--aria2c-x", type=int, default=16, help="Conexiones paralelas por archivo para aria2c (default 16)")
    args = parser.parse_args()

    if not args.xlsx.exists():
        log(f"[ERROR] No existe el archivo: {args.xlsx}")
        return 1

    rows = load_rows(args.xlsx)
    if not rows:
        log("[ERROR] No encontré filas válidas con URL.")
        return 1

    selected = rows[max(args.start_row - 1, 0):]
    if args.limit > 0:
        selected = selected[: args.limit]

    log(f"[INFO] Registros a procesar: {len(selected)}")
    log("[INFO] Usa solo este script sobre grabaciones a las que tengas acceso autorizado.")

    with sync_playwright() as playwright:
        context = resolve_browser_context(playwright, args.headless, args.output)
        try:
            for idx, item in enumerate(selected, start=args.start_row):
                try:
                    process_item(context, item, idx, args.output)
                except Exception as exc:
                    log(f"[ERROR] Fila {item['row']}: {exc}")
        finally:
            context.close()

    log("\n[INFO] Proceso terminado.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
